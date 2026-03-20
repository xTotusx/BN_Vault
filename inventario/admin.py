from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as ReportLabImage
from reportlab.lib.units import inch

from .models import Equipo, EquipoImagen, Pallet, Recepcion

# =========================================================
# FUNCIÓN MAESTRA PARA ESTILOS DE EXCEL (AZUL)
# =========================================================
def aplicar_estilo_excel(ws):
    blue_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row_idx, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            if row_idx == 0:  
                cell.fill = blue_fill 
                cell.font = white_font 

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width if adjusted_width > 12 else 12


# =========================================================
# PANEL 1: RECEPCIONES (¡AHORA CON REGISTRO DE SERIES!)
# =========================================================
# 1. Creamos la mini-tabla para que el inventariador registre las series aquí
class EquipoRecepcionInline(admin.TabularInline):
    model = Equipo
    # Solo mostramos lo que le importa al inventariador
    fields = ['tipo_item', 'marca', 'equipo', 'serie']
    extra = 1 # Muestra 1 fila vacía por defecto (puedes agregar más con un botón)

@admin.register(Recepcion)
class RecepcionAdmin(admin.ModelAdmin):
    list_display = ('guia_rastreo', 'proyecto', 'origen', 'fecha_recepcion', 'equipos_procesados')
    search_fields = ('guia_rastreo', 'proyecto', 'origen')
    list_filter = ('fecha_recepcion', 'proyecto')
    
    # 2. Conectamos la mini-tabla a la vista de Recepciones
    inlines = [EquipoRecepcionInline]

    def equipos_procesados(self, obj):
        return obj.equipos.count()
    equipos_procesados.short_description = "Artículos Escaneados"


# =========================================================
# PANEL 2: EQUIPOS (RESTRINGIDO PARA PRACTICANTES)
# =========================================================
class EquipoImagenInline(admin.TabularInline):
    model = EquipoImagen
    extra = 1 
    fields = ['imagen', 'descripcion']

@admin.register(Equipo)
class EquipoAdmin(admin.ModelAdmin):
    class Media:
        js = ('js/dinamico.js',)

    list_display = ('obtener_proyecto', 'historial_ingresos', 'tipo_item', 'equipo', 'serie', 'estatus', 'inge', 'fecha')
    list_filter = ('tipo_item', 'recepcion__proyecto', 'estatus', 'inge')
    search_fields = ('serie', 'diagnostico', 'marca', 'recepcion__proyecto', 'recepcion__guia_rastreo')
    actions = ["exportar_a_excel", "generar_reporte_pdf"]
    
    inlines = [EquipoImagenInline]
    
    # ¡LA MAGIA DE SEGURIDAD! Excluimos la guía para que nadie pueda cambiarla desde aquí
    exclude = ('recepcion',)

    # Quitamos 'recepcion' del panel visual
    fieldsets = (
        ('Información General', {'fields': ('tipo_item', 'equipo', 'marca', 'serie', 'serie_remplazo')}),
        ('Diagnóstico', {'fields': ('diagnostico', 'estatus', 'inge')}),
        ('Componentes CPU', {'classes': ('grupo-cpu',), 'fields': ('fuente', 'ventilador', 'ssd', 'extensor', 'gabinete', 'disipador', 'mb', 'memoria_ram', 'adaptador_red')}),
        ('Series de Componentes CPU', {'classes': ('grupo-cpu',), 'fields': ('n_serie_fuente', 'n_serie_mb', 'n_serie_ram', 'n_serie_ssd', 'n_serie_gabinete')}),
        ('Componentes Escáner', {'classes': ('grupo-escaner',), 'fields': ('cable_usb', 'base_escaner', 'placa_interna')}),
        ('Componentes Monitor', {'classes': ('grupo-monitor',), 'fields': ('base_monitor', 'cable_hdmi', 'cable_corriente')}),
        ('Componentes Cajón de Dinero', {'classes': ('grupo-cajon',), 'fields': ('bandeja_interna', 'llave', 'cable_cajon')}),
        ('Componentes Impresora', {'classes': ('grupo-impresora',), 'fields': ('navaja', 'sensor_papel', 'plancha_termica', 'motor', 'placa', 'modulo_boton')}),
    )

    def obtener_proyecto(self, obj):
        return obj.recepcion.proyecto
    obtener_proyecto.short_description = "Proyecto"

    def historial_ingresos(self, obj):
        if not obj.serie or obj.serie.upper() in ["N/A", "NA", "-"]:
            return "Sin Serie"
        cuenta = Equipo.objects.filter(serie=obj.serie).count()
        if cuenta > 1:
            return f"⚠️ {cuenta} Ingresos"
        return "1 (Nuevo)"
    historial_ingresos.short_description = "Historial"

    def exportar_a_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario General"
        headers = ['PROYECTO', 'GUIA ENTRADA', 'TIPO', 'EQUIPO', 'MARCA', 'SERIE', 'DIAGNOSTICO', 'ESTATUS', 'INGE', 'FECHA']
        ws.append(headers)

        for obj in queryset:
            fecha_limpia = obj.fecha.replace(tzinfo=None) if obj.fecha else ""
            ws.append([
                obj.recepcion.proyecto, obj.recepcion.guia_rastreo, obj.get_tipo_item_display(), 
                obj.equipo, obj.marca, obj.serie, obj.diagnostico, obj.estatus, obj.inge, fecha_limpia
            ])
            
        aplicar_estilo_excel(ws)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="reporte_inventario_general.xlsx"'
        wb.save(response)
        return response
    exportar_a_excel.short_description = "Generar Excel de seleccionados"

    def generar_reporte_pdf(self, request, queryset):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(name='Title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=16, alignment=1, spaceAfter=20)
        style_heading = ParagraphStyle(name='Heading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, spaceAfter=10, spaceBefore=10)
        style_normal = styles['Normal']
        
        story = []

        for idx, obj in enumerate(queryset):
            story.append(Paragraph("REPORTE INDIVIDUAL - BN VAULT", style_title))
            fecha_str = obj.fecha.strftime('%d/%m/%Y') if obj.fecha else 'N/A'
            story.append(Paragraph(f"<b>Fecha:</b> {fecha_str}", style_normal))
            story.append(Paragraph(f"<b>Ingeniero Asignado:</b> {obj.inge}", style_normal))
            story.append(Spacer(1, 0.2*inch))

            data_equipo = [
                ["Guía / Proyecto:", f"{obj.recepcion.guia_rastreo} ({obj.recepcion.proyecto})", "Tipo:", obj.get_tipo_item_display()],
                ["Equipo:", obj.equipo, "Marca:", obj.marca],
                ["Serie Principal:", obj.serie, "Serie Remplazo:", obj.serie_remplazo if obj.serie_remplazo else "N/A"],
                ["Estatus:", obj.estatus, "Historial:", self.historial_ingresos(obj)]
            ]
            t_equipo = Table(data_equipo, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.8*inch])
            t_equipo.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'), ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
                ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'), ('FONTNAME', (1,0), (1,-1), 'Helvetica'),
                ('FONTNAME', (3,0), (3,-1), 'Helvetica'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(t_equipo)
            story.append(Spacer(1, 0.3*inch))

            story.append(Paragraph("Checklist de Componentes:", style_heading))
            data_checklist = [["Estado", "Componente", "Serie (si aplica)"]]
            componentes = []

            if obj.tipo_item == 'CPU':
                componentes = [
                    (obj.fuente, obj._meta.get_field('fuente').verbose_name, obj.n_serie_fuente),
                    (obj.mb, obj._meta.get_field('mb').verbose_name, obj.n_serie_mb),
                    (obj.memoria_ram, obj._meta.get_field('memoria_ram').verbose_name, obj.n_serie_ram),
                    (obj.ssd, obj._meta.get_field('ssd').verbose_name, obj.n_serie_ssd),
                    (obj.gabinete, obj._meta.get_field('gabinete').verbose_name, obj.n_serie_gabinete),
                    (obj.ventilador, obj._meta.get_field('ventilador').verbose_name, ""),
                    (obj.disipador, obj._meta.get_field('disipador').verbose_name, ""),
                    (obj.extensor, obj._meta.get_field('extensor').verbose_name, ""),
                    (obj.adaptador_red, obj._meta.get_field('adaptador_red').verbose_name, "")
                ]
            elif obj.tipo_item == 'ESCANER':
                componentes = [
                    (obj.cable_usb, obj._meta.get_field('cable_usb').verbose_name, ""),
                    (obj.base_escaner, obj._meta.get_field('base_escaner').verbose_name, ""),
                    (obj.placa_interna, obj._meta.get_field('placa_interna').verbose_name, ""),
                ]
            elif obj.tipo_item == 'MONITOR':
                componentes = [
                    (obj.base_monitor, obj._meta.get_field('base_monitor').verbose_name, ""),
                    (obj.cable_hdmi, obj._meta.get_field('cable_hdmi').verbose_name, ""),
                    (obj.cable_corriente, obj._meta.get_field('cable_corriente').verbose_name, ""),
                ]
            elif obj.tipo_item == 'CAJON':
                componentes = [
                    (obj.bandeja_interna, obj._meta.get_field('bandeja_interna').verbose_name, ""),
                    (obj.llave, obj._meta.get_field('llave').verbose_name, ""),
                    (obj.cable_cajon, obj._meta.get_field('cable_cajon').verbose_name, ""),
                ]
            elif obj.tipo_item == 'IMPRESORA':
                componentes = [
                    (obj.navaja, obj._meta.get_field('navaja').verbose_name, ""),
                    (obj.sensor_papel, obj._meta.get_field('sensor_papel').verbose_name, ""),
                    (obj.plancha_termica, obj._meta.get_field('plancha_termica').verbose_name, ""),
                    (obj.motor, obj._meta.get_field('motor').verbose_name, ""),
                    (obj.placa, obj._meta.get_field('placa').verbose_name, ""),
                    (obj.modulo_boton, obj._meta.get_field('modulo_boton').verbose_name, ""),
                ]
            else:
                componentes = [(False, "No aplica componentes específicos", "")]
            
            for tiene, nombre, serie in componentes:
                estado = "[ X ]" if tiene else "[   ]"
                data_checklist.append([estado, nombre, serie])
                
            t_checklist = Table(data_checklist, colWidths=[1*inch, 2*inch, 3*inch])
            t_checklist.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey), ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
            ]))
            story.append(t_checklist)
            story.append(Spacer(1, 0.3*inch))
            
            story.append(Paragraph("Diagnóstico / Notas:", style_heading))
            story.append(Paragraph(obj.diagnostico, style_normal))
            story.append(Spacer(1, 0.4*inch))

            story.append(Paragraph("Evidencias del Equipo Subidas al Sistema:", style_heading))
            imagenes = obj.imagenes.all()
            if not imagenes:
                story.append(Paragraph("<i>(No se encontraron imágenes de evidencia)</i>", style_normal))
            else:
                for img_obj in imagenes:
                    img_path = img_obj.imagen.path
                    try:
                        rimg = ReportLabImage(img_path, width=6*inch, height=4*inch, kind='proportional')
                        rimg.hAlign = 'CENTER'
                        story.append(rimg)
                        story.append(Spacer(1, 0.1*inch))
                        if img_obj.descripcion:
                            story.append(Paragraph(f"<i>Descripción: {img_obj.descripcion}</i>", ParagraphStyle(name='ImgDesc', parent=styles['Normal'], alignment=1)))
                        story.append(Spacer(1, 0.3*inch))
                    except Exception as e:
                        story.append(Paragraph(f"<b>Error cargando imagen:</b> {img_path}", ParagraphStyle(name='ErrImg', parent=styles['Normal'], textColor=colors.red)))

            if idx < queryset.count() - 1:
                from reportlab.platypus import PageBreak
                story.append(PageBreak())

        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_individual.pdf"'
        return response
    generar_reporte_pdf.short_description = "Generar Reporte Individual PDF con Evidencias Reales"


# =========================================================
# PANEL 3: PALLETS 
# =========================================================
@admin.register(Pallet)
class PalletAdmin(admin.ModelAdmin):
    list_display = ('folio', 'destino', 'fecha_envio', 'total_equipos')
    search_fields = ('folio', 'destino')
    list_filter = ('destino', 'fecha_envio')
    filter_horizontal = ('equipos',) 
    
    actions = ["generar_corte_mensual", "generar_avance_semanal", "generar_pdf_relacion"]

    def total_equipos(self, obj):
        return obj.equipos.count()
    total_equipos.short_description = "Cant. de Artículos"

    def generar_corte_mensual(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Corte Mensual"
        
        headers = [
            'GUIA RETORNO', 'FECHA', 'LOCALIDAD', 'EQUIPO', 'MARCA', 
            'SERIE ORIGINAL', 'SERIE REEMPLAZO', 'DIAGNOSTICO', 'STATUS', 
            'FUENTE', 'VENTILADOR', 'SSD', 'EXTENSOR', 'GABINETE', 
            'DISIPADOR', 'MOTHER', 'RAM', 'ADAPTADOR DE RED', 'PROYECTO', 'GUIA ENTRADA'
        ]
        ws.append(headers)

        for pallet in queryset:
            for eq in pallet.equipos.all():
                ws.append([
                    pallet.folio,
                    pallet.fecha_envio.strftime('%d/%m/%y'),
                    pallet.destino,
                    eq.tipo_item,
                    eq.marca,
                    eq.serie,
                    eq.serie_remplazo if eq.serie_remplazo else "",
                    eq.diagnostico,
                    eq.estatus,
                    "SI" if eq.fuente else "NO",
                    "SI" if eq.ventilador else "NO",
                    "SI" if eq.ssd else "NO",
                    "SI" if eq.extensor else "NO",
                    "SI" if eq.gabinete else "NO",
                    "SI" if eq.disipador else "NO",
                    "SI" if eq.mb else "NO",
                    "SI" if eq.memoria_ram else "NO",
                    "SI" if eq.adaptador_red else "NO",
                    eq.recepcion.proyecto,
                    eq.recepcion.guia_rastreo
                ])

        aplicar_estilo_excel(ws)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="Corte_Mensual.xlsx"'
        wb.save(response)
        return response
    generar_corte_mensual.short_description = "Generar Corte Mensual Excel (Con Guías)"

    def generar_avance_semanal(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Avance Semanal"
        
        headers = [
            'GUIA RETORNO', 'FECHA', 'LOCALIDAD', 'EQUIPO', 'MARCA', 
            'SERIE ORIGINAL', 'SERIE REEMPLAZO', 'DIAGNOSTICO', 'STATUS', 
            'FUENTE', 'VENTILADOR', 'SSD', 'EXTENSOR', 'GABINETE', 
            'DISIPADOR', 'MOTHER', 'RAM', 'ADAPTADOR DE RED', 'PROYECTO'
        ]
        ws.append(headers)

        for pallet in queryset:
            for eq in pallet.equipos.all():
                ws.append([
                    pallet.folio,
                    pallet.fecha_envio.strftime('%d/%m/%y'),
                    pallet.destino,
                    eq.tipo_item,
                    eq.marca,
                    eq.serie,
                    eq.serie_remplazo if eq.serie_remplazo else "",
                    eq.diagnostico,
                    eq.estatus,
                    "SI" if eq.fuente else "NO",
                    "SI" if eq.ventilador else "NO",
                    "SI" if eq.ssd else "NO",
                    "SI" if eq.extensor else "NO",
                    "SI" if eq.gabinete else "NO",
                    "SI" if eq.disipador else "NO",
                    "SI" if eq.mb else "NO",
                    "SI" if eq.memoria_ram else "NO",
                    "SI" if eq.adaptador_red else "NO",
                    eq.recepcion.proyecto
                ])

        aplicar_estilo_excel(ws)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="Avance_Semanal.xlsx"'
        wb.save(response)
        return response
    generar_avance_semanal.short_description = "Generar Avance Semanal Excel (Sin Guías)"

    def generar_pdf_relacion(self, request, queryset):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=50)
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(name='Title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=16, alignment=1, spaceAfter=20)
        style_heading = ParagraphStyle(name='Heading', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12, spaceAfter=10, spaceBefore=10)
        style_normal = styles['Normal']
        
        story = []

        for idx, pallet in enumerate(queryset):
            story.append(Paragraph(f"MANIFIESTO DE ENVÍO - {pallet.folio}", style_title))
            fecha_str = pallet.fecha_envio.strftime('%d/%m/%Y') if pallet.fecha_envio else 'N/A'
            story.append(Paragraph(f"<b>Destino:</b> {pallet.destino}", style_normal))
            story.append(Paragraph(f"<b>Fecha de Envío:</b> {fecha_str}", style_normal))
            story.append(Paragraph(f"<b>Total de Artículos:</b> {pallet.equipos.count()}", style_normal))
            story.append(Spacer(1, 0.3*inch))

            story.append(Paragraph("Relación de Artículos en el Pallet:", style_heading))
            data_equipos = [["#", "Tipo", "Marca", "Serie Original", "Serie Reemplazo", "Proyecto"]]
            
            for i, eq in enumerate(pallet.equipos.all(), 1):
                data_equipos.append([
                    str(i), eq.tipo_item, eq.marca, eq.serie,
                    eq.serie_remplazo if eq.serie_remplazo else "N/A", eq.recepcion.proyecto
                ])
                
            t_equipos = Table(data_equipos, colWidths=[0.3*inch, 1.0*inch, 1.2*inch, 1.5*inch, 1.5*inch, 2.0*inch])
            t_equipos.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('FONTSIZE', (0,0), (-1,-1), 9), 
            ]))
            story.append(t_equipos)
            story.append(Spacer(1, 0.8*inch))
            
            data_firmas = [
                ["___________________________", "", "___________________________"],
                ["Entregó (Firma y Nombre)", "", "Recibió (Firma y Nombre)"]
            ]
            t_firmas = Table(data_firmas, colWidths=[2.5*inch, 2.0*inch, 2.5*inch])
            t_firmas.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ]))
            story.append(t_firmas)

            if idx < queryset.count() - 1:
                from reportlab.platypus import PageBreak
                story.append(PageBreak())

        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relacion_envio_pallets.pdf"'
        return response
    generar_pdf_relacion.short_description = "Generar PDF de Relación de Envío"